# excel_render.py

'''
FileName:	excel_render.py
CreateDate: Aug 10, 2020

Python Library required: 
	- 'pip install openpyxl'
	- 'pip install pillow' 

Readme: This program will render output file into excel format

Execute Command:	python excel_render.py
'''

import os
import sys
from enum import Enum

from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Font, Border, Side, NamedStyle, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook


class Color(Enum):
	RED = 'ff7373'
	YELLOW = 'fff68f'
	GREEN = 'a0db8e'
	BLACK = '000000'
	YELLOW_LINE = 'ffd700'
	TITLE = 'bfefff'
	BACKGROUND = 'fff4db'


# main function
def main():
	input_file_dir = '../output/output.txt'
	if os.path.exists(input_file_dir)==False:
		print("Warning: File {} not exists, please check!".format(input_file_dir))
		sys.exit()


	output_file_name = 'Agent_Status.xlsx'


	# Title Style
	title_style = NamedStyle(name="title_style")
	bd = Side(style='thin', color='000000')
	title_style.border = Border(left=bd, right=bd, top=bd, bottom=bd)
	title_style.fill = PatternFill("solid", fgColor=Color.TITLE.value)
	title_style.font = Font(bold=True)
	title_style.alignment = Alignment(horizontal="center", vertical="center")

	# Running Style
	running_style = NamedStyle(name="running_style")
	bd = Side(style='thin', color='000000')
	running_style.border = Border(left=bd, right=bd, top=bd, bottom=bd)
	running_style.fill = PatternFill("solid", fgColor=Color.GREEN.value)
	running_style.font = Font(bold=True)
	running_style.alignment = Alignment(horizontal="center", vertical="center")

	# Not Running Style
	not_running_style = NamedStyle(name="not_running_style")
	bd = Side(style='thin', color='000000')
	not_running_style.border = Border(left=bd, right=bd, top=bd, bottom=bd)
	not_running_style.fill = PatternFill("solid", fgColor=Color.YELLOW.value)
	not_running_style.font = Font(bold=True)
	not_running_style.alignment = Alignment(horizontal="center", vertical="center")

	# Not Installed Style
	not_installed_style = NamedStyle(name="not_installed_style")
	bd = Side(style='thin', color='000000')
	not_installed_style.border = Border(left=bd, right=bd, top=bd, bottom=bd)
	not_installed_style.fill = PatternFill("solid", fgColor=Color.RED.value)
	not_installed_style.font = Font(bold=True)
	not_installed_style.alignment = Alignment(horizontal="center", vertical="center")


	# Create new excel sheet
	wb_output = Workbook()
	wb_output.remove(wb_output["Sheet"])

	# Add all style into workbook
	wb_output.add_named_style(title_style)
	wb_output.add_named_style(running_style)
	wb_output.add_named_style(not_running_style)
	wb_output.add_named_style(not_installed_style)

	# Create Cover sheet
	ws_output_cover = wb_output.create_sheet('Qualys Agent')
	img = Image('../resources/logo.jpg')
	img.width = 1800
	img.height = 920
	ws_output_cover.add_image(img, 'D2')


	# Create multi sheet: Intalled - Running, Installed - NotRunning , NotInstall
	ws_output_running = wb_output.create_sheet('Running')

	ws_output_running.append(['Hostname', 'Version'])

	ws_output_not_running = wb_output.create_sheet('Not Running')
	ws_output_not_running.append(['Hostname', 'Version'])

	ws_output_not_installed = wb_output.create_sheet('Not Installed')
	ws_output_not_installed.append(['Hostname', 'Version'])


	# Read input_file render into excel sheet
	with open(input_file_dir) as input_file:
		for cnt, line in enumerate(input_file):
			line_array = line.split(',')

			if line_array[1]=='running':
				ws_output_running.append([line_array[0], line_array[2]])
			elif line_array[1]=='not_running' and 'NA' in line_array[2]:
				ws_output_not_installed.append([line_array[0], line_array[2]])
			else:
				ws_output_not_running.append([line_array[0], line_array[2]])


	# adding style into Title
	for i in range(ws_output_running.max_column):
		ws_output_running[get_column_letter(i + 1) + str(1)].style = 'title_style'
		ws_output_running.column_dimensions[get_column_letter(i + 1)].width = 40

	for i in range(ws_output_not_running.max_column):
		ws_output_not_running[get_column_letter(i + 1) + str(1)].style = 'title_style'
		ws_output_not_running.column_dimensions[get_column_letter(i + 1)].width = 40

	for i in range(ws_output_not_installed.max_column):
		ws_output_not_installed[get_column_letter(i + 1) + str(1)].style = 'title_style'
		ws_output_not_installed.column_dimensions[get_column_letter(i + 1)].width = 40


	# adding style into Body
	for i in range(ws_output_running.max_column):
			for j in range(1, ws_output_running.max_row):
				ws_output_running['%s%d' % (get_column_letter(i + 1), j+1)].style = 'running_style'

	for i in range(ws_output_not_running.max_column):
			for j in range(1, ws_output_not_running.max_row):
				ws_output_not_running['%s%d' % (get_column_letter(i + 1), j+1)].style = 'not_running_style'

	for i in range(ws_output_not_installed.max_column):
			for j in range(1, ws_output_not_installed.max_row):
				ws_output_not_installed['%s%d' % (get_column_letter(i + 1), j+1)].style = 'not_installed_style'

	wb_output.save('../excel/'+ output_file_name)


# Run the main method
if __name__ == "__main__":
	main()