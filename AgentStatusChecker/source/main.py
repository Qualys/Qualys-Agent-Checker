import os
import subprocess
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles.borders import Border, Side

# import pandas
base_path = Path(__file__).parent
output_file_path = (base_path / "../output/final_output.txt").resolve()
check_py_file_path = (base_path / "../source/check.py").resolve()
thin_border = Border(bottom=Side(style='thin'))

fmt = "%Y-%m-%d %H:%M:%S %Z%z"

yml_file_path = str((base_path / "../resources/check.yml").resolve())
inv_file_path = str((base_path / "../resources/inv.yml").resolve())
tmp_output_path = str((base_path / "../output/tmp.txt").resolve())
txt_output_path = str((base_path / "../output/output.txt").resolve())
vault_pass_path = str((base_path / "../resources/password_file").resolve())
excel_render_path = str((base_path / "../source/excel_render.py").resolve())

template = '''- hosts: all
  gather_facts: False
  become: True
  become_method: sudo
  ignore_errors: no
  tasks:
   - name: Checking cloud agent service
#     async: 200
#     poll: 200
     script: {CHECK_PYTHON_PATH}
     register: output

   - debug:
       msg: "{{ output.stdout_lines }}" '''
template = template.replace('{CHECK_PYTHON_PATH}', str(check_py_file_path))
open(yml_file_path, 'w').write(template)
os.system('rm -rf /root/.ssh/known_hosts')
l_password = ""
l_user = ""
log = ""
wb = Workbook()
final_output = ''
final_output_file = open(output_file_path, 'w')
txt_output = ''

my_green = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=openpyxl.styles.colors.Color(rgb='0000CC00'))
my_red = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=openpyxl.styles.colors.Color(rgb='00B72C24'))
my_yellow = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=openpyxl.styles.colors.Color(rgb='00E6BC31'))
my_gray = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=openpyxl.styles.colors.Color(rgb='00484466'))
cloud_agent_output = ''


def main():
    global l_user
    global l_password
    global log
    global final_output
    global now
    global cloud_agent_output
    global txt_output
    l_user = ''
    l_password = ''

    # l_user = input('Please input your user name:')
    f1 = open(inv_file_path, 'r').readlines()
    l_user = f1[2].replace(' ','').replace('ansible_user:', '').strip()
    if l_user.replace(' ', '') == '':
        print('Please input your info into resources/inv.yml')
        return

    # except:
    #     l_user = input('Please input your user name:')
    # l_password = getpass.getpass("password: ")

    os.system('export ANSIBLE_HOST_KEY_CHECKING=False')
    os.putenv('ANSIBLE_HOST_KEY_CHECKING', 'FALSE')

    final_output = final_output + '['

    check_hosts()

    final_output = final_output + ']'

    final_output_file.write(final_output)

    final_output_file.close()

    colored_wb = Workbook()

    colored_ws = colored_wb.active

    colored_idx = 2
    f = open(output_file_path).readlines()

    for line in f:
        if 'hostname' not in line:
            continue
        line = line.replace('[', '')
        line = line.replace(']', '')
        splited = line.split(',')
        cloud_agent_output = cloud_agent_output + splited[0].replace('"', '').replace(':', '') \
            .replace('hostname', '').replace('{', '') + ',' + splited[1].replace('"', '').replace(
            'cloud_agent_version', '').replace(':', '').replace('}', '') + ',' + splited[2].replace('"', '').replace(
            'cloud_agent', '').replace(':', '').replace('}', '') + '\n'
        txt_output = txt_output + splited[0].replace('"', '').replace(':', '') \
            .replace('hostname', '').replace('{', '') + ',' + splited[2].replace('"', '').replace('cloud_agent', '') \
                         .replace(':', '').replace('}', '') + ',' + splited[1].replace('"', '').replace(
            'cloud_agent_version', '') \
                         .replace(':', '').replace('}', '') + '\n'

        # colored_ws.cell(row=colored_idx, column=1).value = splited[0].replace('"', '').replace(':', '') \
        #    .replace('hostname', '').replace('{', '')

        # colored_ws.cell(row=colored_idx, column=2).value = splited[1].replace('"', '').replace('cloud_agent_version',
        #                                                                                        '') \
        #     .replace(':', '').replace('}', '')
        #
        # colored_ws.cell(row=colored_idx, column=3).value = splited[2].replace('"', '').replace('cloud_agent', '') \
        #     .replace(':', '').replace('}', '')
        #
        # if 'not_running' in line:
        #     colored_ws.cell(row=colored_idx, column=3).fill = my_red
        # else:
        #     colored_ws.cell(row=colored_idx, column=3).fill = my_green
        #
        # if 'not_installed' in line:
        #     colored_ws.cell(row=colored_idx, column=2).fill = my_red
        #
        # else:
        #     colored_ws.cell(row=colored_idx, column=2).fill = my_green
        # colored_idx = colored_idx + 1

    # for column_cells in colored_ws.columns:
    #     length = max(len(str(cell.value)) + 3 for cell in column_cells)
    #     colored_ws.column_dimensions[column_cells[0].column_letter].width = length
    #
    # colored_ws.cell(row=1, column=1).value = 'hostname ##Update time: ' + now
    #
    # colored_ws.cell(row=1, column=2).value = 'qualys_cloud_agent_version '
    #
    # colored_ws.cell(row=1, column=3).value = 'qualys_cloud_agent '
    # xlsx_path = str((base_path / "../output/final_output.xlsx").resolve())
    open(txt_output_path, 'w').write(txt_output)
    # colored_wb.save(xlsx_path)
    # os.system('soffice --headless --convert-to html ' + xlsx_path)
    # os.system('soffice --headless --convert-to html ' + xlsx_path)
    excel_render()


def check_hosts():
    global log
    global wb
    global final_output

    output_data = ''

    print(('ansible-playbook %s  -i  %s -f 20 --vault-id %s@%s'
           % (yml_file_path, inv_file_path, l_user, vault_pass_path)))
    os.system('ansible-playbook %s  -i  %s -f 20 --vault-id %s@%s 2>&1 | tee %s'
              % (yml_file_path, inv_file_path, l_user, vault_pass_path, tmp_output_path))
    lines = open(tmp_output_path).readlines()
    start = 0
    b = False
    out = ''
    for i in range(0, len(lines)):
        if 'TASK [debug] *************************************' in lines[i]:
            b = True
            continue
        if 'PLAY RECAP ***************************************' in lines[i]:
            b = False
        if b:
            out = out + lines[i]
        if 'PLAY RECAP' in lines[i]:
            start = i
    out_list = out.split('ok: ')

    for l in out_list:
        for e in l.split('\n'):
            if 'hostname' in e:
                e = e.replace('\\', '').replace(' ', '')
                e = e[1: len(e.replace('\\', '')) - 1] + ',\n'
                final_output = final_output + e

    for i in range(start + 1, len(lines)):
        if lines[i].split(' ')[0].replace(' ', '') == '':
            continue

        else:
            output_data = output_data + lines[i].split(' ')[0] + ',succeed,' + '\n'


def send_email(recipient, subject, body):
    try:
        process = subprocess.Popen(['mail', '-s', subject, recipient], stdin=subprocess.PIPE)
        print('report email sent.')
        process.communicate(body)
    except Exception as error:
        print(error)


def excel_render():
    import os
    import sys
    from enum import Enum

    from openpyxl.drawing.image import Image
    from openpyxl.styles import PatternFill, Font, Border, Side, NamedStyle, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.workbook import Workbook

    class Color(Enum):
        RED = 'ff7373'
        YELLOW = 'fff68f'
        GREEN = 'a0db8e'
        BLACK = '000000'
        YELLOW_LINE = 'ffd700'
        TITLE = 'bfefff'
        BACKGROUND = 'fff4db'

    # main function
    def main_render():
        input_file_dir = '../output/output.txt'
        if os.path.exists(input_file_dir) == False:
            print("Warning: File {} not exists, please check!".format(input_file_dir))
            sys.exit()

        output_file_name = 'Agent_Status.xlsx'

        # Title Style
        title_style = NamedStyle(name="title_style")
        bd = Side(style='thin', color='000000')
        title_style.border = Border(left=bd, right=bd, top=bd, bottom=bd)
        title_style.fill = PatternFill("solid", fgColor=Color.TITLE.value)
        title_style.font = Font(bold=True)
        title_style.alignment = Alignment(horizontal="center", vertical="center")

        # Running Style
        running_style = NamedStyle(name="running_style")
        bd = Side(style='thin', color='000000')
        running_style.border = Border(left=bd, right=bd, top=bd, bottom=bd)
        running_style.fill = PatternFill("solid", fgColor=Color.GREEN.value)
        running_style.font = Font(bold=True)
        running_style.alignment = Alignment(horizontal="center", vertical="center")

        # Not Running Style
        not_running_style = NamedStyle(name="not_running_style")
        bd = Side(style='thin', color='000000')
        not_running_style.border = Border(left=bd, right=bd, top=bd, bottom=bd)
        not_running_style.fill = PatternFill("solid", fgColor=Color.YELLOW.value)
        not_running_style.font = Font(bold=True)
        not_running_style.alignment = Alignment(horizontal="center", vertical="center")

        # Not Installed Style
        not_installed_style = NamedStyle(name="not_installed_style")
        bd = Side(style='thin', color='000000')
        not_installed_style.border = Border(left=bd, right=bd, top=bd, bottom=bd)
        not_installed_style.fill = PatternFill("solid", fgColor=Color.RED.value)
        not_installed_style.font = Font(bold=True)
        not_installed_style.alignment = Alignment(horizontal="center", vertical="center")

        # Create new excel sheet
        wb_output = Workbook()
        wb_output.remove(wb_output["Sheet"])

        # Add all style into workbook
        wb_output.add_named_style(title_style)
        wb_output.add_named_style(running_style)
        wb_output.add_named_style(not_running_style)
        wb_output.add_named_style(not_installed_style)

        # Create Cover sheet
        ws_output_cover = wb_output.create_sheet('Qualys Agent')
        img = Image('../resources/logo.jpg')
        img.width = 1800
        img.height = 920
        ws_output_cover.add_image(img, 'D2')

        # Create multi sheet: Intalled - Running, Installed - NotRunning , NotInstall
        ws_output_running = wb_output.create_sheet('Running')

        ws_output_running.append(['Hostname', 'Version'])

        ws_output_not_running = wb_output.create_sheet('Not Running')
        ws_output_not_running.append(['Hostname', 'Version'])

        ws_output_not_installed = wb_output.create_sheet('Not Installed')
        ws_output_not_installed.append(['Hostname', 'Version'])

        # Read input_file render into excel sheet
        with open(input_file_dir) as input_file:
            for cnt, line in enumerate(input_file):
                line_array = line.split(',')

                if line_array[1] == 'running':
                    ws_output_running.append([line_array[0], line_array[2]])
                elif line_array[1] == 'not_running' and 'NA' in line_array[2]:
                    ws_output_not_installed.append([line_array[0], line_array[2]])
                else:
                    ws_output_not_running.append([line_array[0], line_array[2]])

        # adding style into Title
        for i in range(ws_output_running.max_column):
            ws_output_running[get_column_letter(i + 1) + str(1)].style = 'title_style'
            ws_output_running.column_dimensions[get_column_letter(i + 1)].width = 40

        for i in range(ws_output_not_running.max_column):
            ws_output_not_running[get_column_letter(i + 1) + str(1)].style = 'title_style'
            ws_output_not_running.column_dimensions[get_column_letter(i + 1)].width = 40

        for i in range(ws_output_not_installed.max_column):
            ws_output_not_installed[get_column_letter(i + 1) + str(1)].style = 'title_style'
            ws_output_not_installed.column_dimensions[get_column_letter(i + 1)].width = 40

        # adding style into Body
        for i in range(ws_output_running.max_column):
            for j in range(1, ws_output_running.max_row):
                ws_output_running['%s%d' % (get_column_letter(i + 1), j + 1)].style = 'running_style'

        for i in range(ws_output_not_running.max_column):
            for j in range(1, ws_output_not_running.max_row):
                ws_output_not_running['%s%d' % (get_column_letter(i + 1), j + 1)].style = 'not_running_style'

        for i in range(ws_output_not_installed.max_column):
            for j in range(1, ws_output_not_installed.max_row):
                ws_output_not_installed['%s%d' % (get_column_letter(i + 1), j + 1)].style = 'not_installed_style'

        wb_output.save('../excel/' + output_file_name)

    main_render()


if __name__ == "__main__":
    main()
